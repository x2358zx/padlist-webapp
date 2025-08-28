import os
import io
import uuid
import zipfile
from typing import List, Optional, Dict, Any

from fastapi import FastAPI, Request, UploadFile, File, Form
from fastapi.responses import HTMLResponse, JSONResponse, FileResponse, RedirectResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates

from openpyxl import load_workbook
from openpyxl.utils.cell import coordinate_from_string
from openpyxl.utils import get_column_letter
from PIL import Image

import json
import posixpath as pp
import xml.etree.ElementTree as ET
import re

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
UPLOAD_DIR = os.path.join(BASE_DIR, "uploads")
os.makedirs(UPLOAD_DIR, exist_ok=True)

app = FastAPI()
app.mount("/static", StaticFiles(directory=os.path.join(BASE_DIR, "static")), name="static")
templates = Jinja2Templates(directory=os.path.join(BASE_DIR, "templates"))

# === 身分/環境設定（以環境變數為主） ===
TRUST_PROXY = os.getenv("TRUST_PROXY", "1") == "1"
DEPLOY_ENV = os.getenv("DEPLOY_ENV", "test")
SUPERVISOR_USERS = [u.strip() for u in os.getenv("SUPERVISOR_USERS", "").split(",") if u.strip()]
ALLOWED_EDITOR_IPS = [ip.strip() for ip in os.getenv("ALLOWED_EDITOR_IPS", "").split(",") if ip.strip()]
DEV_ALLOW_LOCAL_EDITOR = os.getenv("DEV_ALLOW_LOCAL_EDITOR", "1") == "1"

def get_client_ip(request: Request) -> str:
    """
    從 X-Forwarded-For / X-Real-IP（在 nginx 設定）或 fallback 到 request.client.host 取得來源 IP
    """
    if TRUST_PROXY:
        xff = request.headers.get("x-forwarded-for")
        if xff:
            # 取最前面那個（最接近客戶端）
            return xff.split(",")[0].strip()
        xri = request.headers.get("x-real-ip")
        if xri:
            return xri.strip()
    # 沒代理或沒標頭時
    return (request.client.host if request.client else "") or ""

def is_editor(request: Request) -> bool:
    """
    伺服端權限判斷（IP / AD 使用者）
    規則：
    1) 若有 AD 使用者（X-Remote-User），且在 SUPERVISOR_USERS 清單中 → 通過
    2) 來源 IP 在 ALLOWED_EDITOR_IPS → 通過
    3) 本機/開發測試：DEV_ALLOW_LOCAL_EDITOR=1 且 IP 為 127.0.0.1 / ::1 → 通過
    """
    ip = get_client_ip(request)
    user = request.headers.get("X-Remote-User") or request.headers.get("Remote-User")

    if user and SUPERVISOR_USERS and user in SUPERVISOR_USERS:
        return True
    if ip and ALLOWED_EDITOR_IPS and ip in ALLOWED_EDITOR_IPS:
        return True
    if DEV_ALLOW_LOCAL_EDITOR and ip in ("127.0.0.1", "::1"):
        return True
    return False

@app.get("/me")
async def me(request: Request):
    """
    回傳前端需要的身分資訊，讓前端決定是否開啟編輯 UI
    """
    ip = get_client_ip(request)
    user = request.headers.get("X-Remote-User") or request.headers.get("Remote-User")
    return JSONResponse({
        "client_ip": ip,
        "user": user,
        "env": DEPLOY_ENV,
        "is_editor": is_editor(request)
    })


@app.get("/favicon.ico")
async def favicon():
    return RedirectResponse(url="/static/app.ico")

# === 新增：將檔名安全化（用於輸出圖檔）===
def _safe_name(name: str) -> str:
    keep = "-_.()[]{}+@！@全形也可用"
    return "".join(ch if ch.isalnum() or ch in keep else "_" for ch in name).strip("_") or "sheet"

# === 新增：建構「每個工作表 → 最大張圖片」對應表，並把圖片解出來到 session 目錄 ===
def _build_sheet_image_map(xlsx_path: str, out_dir: str):
    """
    回傳:
      (ordered_sheet_names, map_name_to_saved_path)
      - ordered_sheet_names: 依 workbook sheets 原始順序的名稱清單
      - map_name_to_saved_path: {sheet_name: /abs/save/path/of/largest_image}
    規則：
      - 只挑每張工作表面積最大的圖片（用 cx*cy 估算）
      - 若該表沒有圖片，略過（不進下拉）
      - 若無尺寸資訊，退回第一張
    """
    with zipfile.ZipFile(xlsx_path, "r") as zf:
        # 1) 解析 workbook.xml，取得 sheet name 與 rid
        wbk_xml = "xl/workbook.xml"
        wbk_rels = "xl/_rels/workbook.xml.rels"
        ns = {
            "main": "http://schemas.openxmlformats.org/spreadsheetml/2006/main",
            "r": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
            "xdr": "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing",
            "a": "http://schemas.openxmlformats.org/drawingml/2006/main",
            "pr": "http://schemas.openxmlformats.org/package/2006/relationships",
        }
        sheets_order = []
        wtree = ET.fromstring(zf.read(wbk_xml))
        for s in wtree.findall(".//main:sheets/main:sheet", ns):
            sheets_order.append((s.attrib.get("name",""), s.attrib.get("{%s}id" % ns["r"], "")))

        # 2) rId → worksheet 路徑
        rid_to_ws = {}
        rtree = ET.fromstring(zf.read(wbk_rels))
        for rel in rtree.findall(".//pr:Relationship", ns):
            rid = rel.attrib.get("Id","")
            tgt = rel.attrib.get("Target","")
            # target 通常像 "worksheets/sheet1.xml"
            rid_to_ws[rid] = pp.normpath(pp.join("xl", tgt))

        # 3) worksheet → drawing → images，挑最大張
        name_to_saved = {}
        for sheet_name, rid in sheets_order:
            ws_path = rid_to_ws.get(rid)
            if not ws_path or ws_path not in zf.namelist():
                continue

            # 找這張工作表的 rels，裡面會有 drawing
            ws_rels = pp.normpath(pp.join("xl/worksheets/_rels", pp.basename(ws_path) + ".rels"))
            if ws_rels not in zf.namelist():
                continue

            wsrels_tree = ET.fromstring(zf.read(ws_rels))
            drawing_target = None
            for rel in wsrels_tree.findall(".//pr:Relationship", ns):
                if rel.attrib.get("Type","").endswith("/drawing"):
                    drawing_target = rel.attrib.get("Target","")  # ex: "../drawings/drawing1.xml"
                    break
            if not drawing_target:
                continue

            drawing_xml = pp.normpath(pp.join(pp.dirname(ws_path), drawing_target))  # → xl/drawings/drawing1.xml
            if drawing_xml not in zf.namelist():
                continue

            # drawing 的 rels：把 r:embed → 影像檔路徑對上
            drawing_rels = pp.normpath(pp.join(pp.dirname(drawing_xml), "_rels", pp.basename(drawing_xml) + ".rels"))
            if drawing_rels not in zf.namelist():
                continue
            drels_tree = ET.fromstring(zf.read(drawing_rels))
            embed_to_media = {}
            for rel in drels_tree.findall(".//pr:Relationship", ns):
                if rel.attrib.get("Type","").endswith("/image"):
                    rid_img = rel.attrib.get("Id","")
                    tgt_img = rel.attrib.get("Target","")  # ex: "../media/image1.png"
                    media_path = pp.normpath(pp.join(pp.dirname(drawing_xml), tgt_img))  # → xl/media/image1.png
                    embed_to_media[rid_img] = media_path

            # 解析 drawing.xml 找出所有圖片的 a:blip（拿 r:embed）與 a:ext（拿 cx, cy）
            dtree = ET.fromstring(zf.read(drawing_xml))
            candidates = []  # [(area, embed_id)]
            for pic in dtree.findall(".//xdr:pic", ns):
                blip = pic.find(".//a:blip", ns)
                if blip is None:
                    continue
                embed_id = blip.attrib.get("{%s}embed" % ns["r"])
                # 嘗試抓尺寸（有些檔可能沒有 ext）
                ext = pic.find(".//a:xfrm/a:ext", ns)
                try:
                    cx = int(ext.attrib.get("cx","0")) if ext is not None else 0
                    cy = int(ext.attrib.get("cy","0")) if ext is not None else 0
                except Exception:
                    cx = cy = 0
                area = cx * cy
                candidates.append((area, embed_id))

            if not candidates:
                continue
            # 依面積挑最大；若都 0，這個排序也會保留第一張
            candidates.sort(key=lambda t: t[0], reverse=True)
            _, best_embed = candidates[0]
            media_rel = embed_to_media.get(best_embed)
            if not media_rel or media_rel not in zf.namelist():
                continue

            # 寫出檔案到 session 目錄
            ext = os.path.splitext(media_rel)[1].lower() or ".png"
            out_name = f"{_safe_name(sheet_name)}_largest{ext}"
            out_path = os.path.join(out_dir, out_name)
            with open(out_path, "wb") as f:
                f.write(zf.read(media_rel))

            name_to_saved[sheet_name] = out_path

        # 只有「有圖」的工作表需要列入選單
        ordered_names_with_image = [nm for nm, _ in sheets_order if nm in name_to_saved]
        return ordered_names_with_image, name_to_saved

def _sheet_has_data(ws) -> bool:
    """Heuristic: if any cell in the used range has a non-empty value."""
    try:
        min_row = 1
        max_row = ws.max_row or 0
        min_col = 1
        max_col = ws.max_column or 0
        max_row = min(max_row, 200)
        max_col = min(max_col, 50)
        for r in range(min_row, max_row + 1):
            for c in range(min_col, max_col + 1):
                if ws.cell(row=r, column=c).value not in (None, ""):
                    return True
        return False
    except Exception:
        return False

def _read_cell_text(ws, cell_addr: str) -> str:
    try:
        value = ws[cell_addr].value
        if value not in (None, ""):
            return str(value)
        # 若空，嘗試從合併儲存格取值
        target_col_letter, target_row = coordinate_from_string(cell_addr)
        target_col_idx = ws[cell_addr].column  # numeric
        for mr in ws.merged_cells.ranges:
            min_col, min_row, max_col, max_row = mr.bounds
            # 與 WPF 相同邏輯：同列、且目標欄在合併範圍內
            if min_row == int(target_row) and min_col <= target_col_idx <= max_col:
                v = ws.cell(row=min_row, column=min_col).value
                if v not in (None, ""):
                    return str(v)
    except Exception:
        pass
    return ""

# === 自動偵測小工具（中文註解） ===
def _scan_text(ws, max_rows=120, max_cols=40):
    """把前 max_rows × max_cols 的儲存格掃一遍，回傳 (r,c,文字) 的生成器。"""
    for r in range(1, max_rows + 1):
        for c in range(1, max_cols + 1):
            v = ws.cell(row=r, column=c).value
            if v not in (None, ""):
                yield r, c, str(v).strip()

def _find_cell(ws, keywords, **kw):
    """在頁面左上角區域找『包含 keywords 任一關鍵字』的儲存格。
    回傳 (row, col)；多筆時採『最靠上、再最靠左』的那一格。"""
    if isinstance(keywords, str):
        keywords = [keywords]
    kws = [k.lower() for k in keywords]
    best = None
    for r, c, s in _scan_text(ws, **kw):
        s_low = s.lower()
        if any(k in s_low for k in kws):
            if best is None or (r < best[0] or (r == best[0] and c < best[1])):
                best = (r, c)
    return best

# 文字正規化：去掉非英數，轉小寫，便於比對「等於」
def _norm(s: str) -> str:
    return re.sub(r'[^a-z0-9]+', '', s.lower()) if s else ''

def _find_header_exact(ws, patterns, **kw):
    """
    找『完全等於』其中一個 pattern（比對用 _norm）
    例如：patterns=["pin", "pinno", "pin#"]，就不會把 "Pin Name" 誤判成 "Pin"
    """
    pats = [_norm(p) for p in (patterns if isinstance(patterns, (list, tuple)) else [patterns])]
    best = None
    for r, c, s in _scan_text(ws, **kw):
        if _norm(s) in pats:
            if best is None or (r < best[0] or (r == best[0] and c < best[1])):
                best = (r, c)
    return best


def _col_letter(cidx: int) -> str:
    from openpyxl.utils import get_column_letter
    return get_column_letter(cidx)

def _find_header_exact_in_row(ws, patterns, row, col_from=1, col_to=None):
    """只在 row 這一列找『完全等於 patterns 之一』的表頭。會回傳 (row, col)。"""
    if isinstance(patterns, (list, tuple)):
        pats = [_norm(p) for p in patterns]
    else:
        pats = [_norm(patterns)]
    if row < 1 or row > (ws.max_row or 0):
        return None
    if col_to is None:
        col_to = min(ws.max_column or 0, 100)

    for c in range(max(1, col_from), col_to + 1):
        v = ws.cell(row=row, column=c).value
        if v not in (None, "") and _norm(str(v)) in pats:
            return (row, c)
    return None


def _extract_first_image_from_xlsx(xlsx_path: str, out_dir: str) -> Optional[str]:
    # 直接從 zip 取 xl/media/* 第一張
    try:
        with zipfile.ZipFile(xlsx_path, "r") as zf:
            media_files = [n for n in zf.namelist() if n.startswith("xl/media/") and (n.lower().endswith(".png") or n.lower().endswith(".jpg") or n.lower().endswith(".jpeg"))]
            if not media_files:
                return None
            name = media_files[0]
            data = zf.read(name)
            ext = os.path.splitext(name)[1].lower()
            out_path = os.path.join(out_dir, f"chip_image{ext}")
            with open(out_path, "wb") as f:
                f.write(data)
            return out_path
    except Exception:
        return None

@app.get("/", response_class=HTMLResponse)
async def index(request: Request):
    return templates.TemplateResponse("index.html", {"request": request})

@app.post("/upload")
async def upload_excel(file: UploadFile = File(...)):
    try:
        sid = uuid.uuid4().hex[:10]
        sess_dir = os.path.join(UPLOAD_DIR, sid)
        os.makedirs(sess_dir, exist_ok=True)

        saved_path = os.path.join(sess_dir, "workbook.xlsx")
        with open(saved_path, "wb") as f:
            f.write(await file.read())

        # wb = load_workbook(saved_path, data_only=True)
        # sheets = [ws.title for ws in wb.worksheets if _sheet_has_data(ws)]
        # if not sheets:
        #     sheets = list(wb.sheetnames)
        # 
        # img_path = _extract_first_image_from_xlsx(saved_path, sess_dir)
        # img_url = f"/uploads/{sid}/" + os.path.basename(img_path) if img_path else None
        # 
        # return JSONResponse({"session_id": sid, "sheets": sheets, "image_url": img_url})
        # 讀 workbook，建立「每表最大圖」對應（只列出有圖的工作表）
        wb = load_workbook(saved_path, data_only=True)
        sheets_with_img_ordered, map_name_to_saved = _build_sheet_image_map(saved_path, sess_dir)

        # 將對應表存成 json，給 /sheet_info 使用
        mapping_json = os.path.join(sess_dir, "sheet_images.json")
        with open(mapping_json, "w", encoding="utf-8") as f:
            json.dump({k: os.path.basename(v) for k, v in map_name_to_saved.items()}, f, ensure_ascii=False)

        # 預設顯示第一個有圖的工作表之圖片
        default_image_url = None
        if sheets_with_img_ordered:
            first_sheet = sheets_with_img_ordered[0]
            fname = os.path.basename(map_name_to_saved[first_sheet])
            default_image_url = f"/uploads/{sid}/{fname}"

        return JSONResponse({
            "session_id": sid,
            # 僅包含「有圖」的工作表，前端下拉就不會出現沒圖的表
            "sheets": sheets_with_img_ordered,
            # 初始圖（第一個工作表的「最大張」）
            "default_image_url": default_image_url
        })
    except Exception as e:
        return JSONResponse({"error": f"Failed to read Excel: {type(e).__name__}: {e}"}, status_code=400)

@app.get("/uploads/{sid}/{fname}")
async def serve_upload(sid: str, fname: str):
    path = os.path.join(UPLOAD_DIR, sid, fname)
    if os.path.exists(path):
        return FileResponse(path)
    return JSONResponse({"error": "file not found"}, status_code=404)

@app.post("/sheet_info")
async def sheet_info(
    session_id: str = Form(...),
    sheet_name: str = Form(...),
):
    sess_dir = os.path.join(UPLOAD_DIR, session_id)
    xlsx_path = os.path.join(sess_dir, "workbook.xlsx")
    if not os.path.exists(xlsx_path):
        return JSONResponse({"error": "session not found"}, status_code=404)

    wb = load_workbook(xlsx_path, data_only=True)
    if sheet_name not in wb.sheetnames:
        return JSONResponse({"error": "sheet not found"}, status_code=404)
    ws = wb[sheet_name]

    # === 自動偵測：Chip Size / Project Code / PadWindow / CUP（中文註解） ===
    # 1) Chip Size：找含「chip size」的關鍵字，往右一格讀取文字並解析 "123 um x 456 um"
    cs_pos = _find_cell(ws, ["chip size", "chipsize", "chip-size"])
    width = height = None
    if cs_pos:
        r, c = cs_pos
        cell_txt = _read_cell_text(ws, f"{_col_letter(c+1)}{r}")
        m = re.search(r"(\d+\.?\d*)\s*um\s*[X×x]\s*(\d+\.?\d*)\s*um", str(cell_txt))
        if m:
            width = float(m.group(1))
            height = float(m.group(2))

    # 2) Project Code：找到「Name」關鍵字，往右一格
    proj_pos = _find_cell(ws, ["name"])
    project_code = None
    if proj_pos:
        r, c = proj_pos
        project_code = _read_cell_text(ws, f"{_col_letter(c+1)}{r}") or ""

    # 3) PadWindow / CUP：各自往右一格（可選）
    padwindow = cup = None
    pw_pos = _find_cell(ws, ["padwindow", "pad window"])
    if pw_pos:
        r, c = pw_pos
        padwindow = _read_cell_text(ws, f"{_col_letter(c+1)}{r}") or ""
    cup_pos = _find_cell(ws, ["cup"])
    if cup_pos:
        r, c = cup_pos
        cup = _read_cell_text(ws, f"{_col_letter(c+1)}{r}") or ""


    # 依工作表回傳對應的「最大張圖片」
    mapping_json = os.path.join(sess_dir, "sheet_images.json")
    img_url = None
    if os.path.exists(mapping_json):
        try:
            with open(mapping_json, "r", encoding="utf-8") as f:
                m = json.load(f)  # {sheet_name: filename}
            fname = m.get(sheet_name)
            if fname:
                img_url = f"/uploads/{session_id}/{fname}"
        except Exception:
            img_url = None

    return JSONResponse({
        "chip_size": {"width": width, "height": height},
        "project_code": project_code,
        "image_url": img_url,
        "extras": {"PadWindow": padwindow, "CUP": cup}
    })

@app.post("/parse_pins")
async def parse_pins(
    session_id: str = Form(...),
    sheet_name: str = Form(...),
):
    sess_dir = os.path.join(UPLOAD_DIR, session_id)
    xlsx_path = os.path.join(sess_dir, "workbook.xlsx")
    if not os.path.exists(xlsx_path):
        return JSONResponse({"error": "session not found"}, status_code=404)

    wb = load_workbook(xlsx_path, data_only=True)
    if sheet_name not in wb.sheetnames:
        return JSONResponse({"error": "sheet not found"}, status_code=404)
    ws = wb[sheet_name]
    if ws.max_row is None or ws.max_row == 0:
        return JSONResponse({"valid_pins": [], "invalid_pins": []})

        # === 自動偵測：PIN / Text Name / X-axis / Y-axis 四個欄位置與起始列 ===
    # 容許不同寫法（大小寫/空白/破折號）
    pin_hdr = _find_header_exact(ws, ["pin", "pinno", "pin#", "pinno."])
    name_hdr = _find_header_exact(ws, ["textname", "pinname", "name"])
    x_hdr   = _find_header_exact(ws, ["xaxis", "x-axis", "x"])
    y_hdr   = _find_header_exact(ws, ["yaxis", "y-axis", "y"])

        # === 讓 Name 表頭「靠近 PIN/X/Y 所在的表頭列」 ===
    header_row_guess = max(pin_hdr[0], x_hdr[0], y_hdr[0])  # 多半同列，取最大那列當表頭列

    # 先嘗試：只在這一列找 name 表頭
    name_near = _find_header_exact_in_row(ws, ["textname", "pinname", "name"], header_row_guess)
    if not name_near:
        # 再放寬到 ±2 列
        for dr in ( -1, 1, -2, 2 ):
            cand = _find_header_exact_in_row(ws, ["textname", "pinname", "name"], header_row_guess + dr)
            if cand:
                name_near = cand
                break
    if name_near:
        name_hdr = name_near

    # 若 name 跟 pin 還是在同一欄，優先從「同列表頭、pin 右邊」再找一次
    if name_hdr and pin_hdr and name_hdr[1] == pin_hdr[1]:
        cand = _find_header_exact_in_row(ws, ["textname", "pinname", "name"],
                                         header_row_guess, col_from=pin_hdr[1] + 1)
        if cand:
            name_hdr = cand


    # 若 pin_hdr 與 name_hdr 指到同一格（例如標頭是 "Pin Name"）
    if pin_hdr and name_hdr and pin_hdr == name_hdr:
        hdr_txt = _read_cell_text(ws, f"{_col_letter(pin_hdr[1])}{pin_hdr[0]}")
        if "name" in (hdr_txt or "").lower():
            # 這格應該歸「Name」，重新搜「Pin No」但限定只找 "PIN/PIN NO/PIN#"
            pin_hdr = _find_header_exact(ws, ["pin", "pinno", "pin#", "pinno."])

    if not (pin_hdr and name_hdr and x_hdr and y_hdr):
        return JSONResponse({"valid_pins": [], "invalid_pins": ["未偵測到表頭（PIN/Name/X-axis/Y-axis）"]})
        

    # 取「最靠下的表頭列」+1 作為資料起始列（避免表頭不在同一列的情況）
    start_row = max(pin_hdr[0], name_hdr[0], x_hdr[0], y_hdr[0]) + 1

    from openpyxl.utils import get_column_letter
    col_pin = get_column_letter(pin_hdr[1])
    col_nam = get_column_letter(name_hdr[1])
    col_x   = get_column_letter(x_hdr[1])
    col_y   = get_column_letter(y_hdr[1])

    def read_cell(col_letter: str, row_idx: int) -> str:
        try:
            v = ws[f"{col_letter}{row_idx}"].value
            s = "" if v in (None, "") else str(v)
            # 🆕 去掉 NBSP(\u00A0) 與全形空白(\u3000)，再 strip
            return s.replace("\u00A0", "").replace("\u3000", "").strip()
        except Exception:
            return ""

    valid_pins, invalid_pins = [], []

    r = start_row
    while r <= ws.max_row:
        pin_no   = read_cell(col_pin, r)
        pin_name = read_cell(col_nam, r)
        num = lambda s: re.sub(r"[^0-9.+-]", "", (s or ""))
        x_text = num(read_cell(col_x, r))
        y_text = num(read_cell(col_y, r))


        # 停止條件：四欄都空白 → 結束掃描
        if pin_no == "" and pin_name == "" and x_text == "" and y_text == "":
            break

        # === 決定這列的「身分」 ===
        has_id = (pin_no != "" or pin_name != "")   # 🆕 只要 PIN 或 NAME 有其一
        has_xy = (x_text != "" or y_text != "")

        # 兩欄都空白，但座標有東西 → 視為雜訊列，直接跳過
        if not has_id and has_xy:
            r += 1
            continue

        # 嘗試把座標轉 float
        try:
            x = float(x_text)
            y = float(y_text)
        except Exception:
            x = y = None

        # 強化 NC 偵測（N/C、n c…都抓得到）
        norm_name = re.sub(r'[^a-z]', '', (pin_name or '').lower())
        is_nc = (norm_name == "nc")

        # === 加入 invalid 或 valid 的規則 ===
        if (pin_no == "" or is_nc or x is None or y is None):
            # 🆕 只有當「至少有 PIN 或 NAME 其中一個」才列入 invalid_pins
            if has_id:
                invalid_pins.append(f"{pin_no}, {(pin_name or '').strip()}")
        else:
            cleaned_name = (pin_name or "").replace(" ", "")
            valid_pins.append({"pin_no": pin_no, "pin_name": cleaned_name, "x": x, "y": y})

        r += 1

    return JSONResponse({"valid_pins": valid_pins, "invalid_pins": invalid_pins})

# === 注意事項（operation / bonding）讀寫 ===
DEFAULT_NOTES = {
    "operation": "（預設）這裡放操作注意事項的說明，供主管編輯…",
    "bonding": "（預設）這裡放 bonding 注意事項的說明，供主管編輯…"
}

@app.get("/notices")
async def get_notices(session_id: str):
    """讀取某次上傳(session)的注意事項；若尚未建立則回預設。"""
    sess_dir = os.path.join(UPLOAD_DIR, session_id)
    path = os.path.join(sess_dir, "notices.json")
    data = DEFAULT_NOTES.copy()
    if os.path.exists(path):
        try:
            with open(path, "r", encoding="utf-8") as f:
                saved = json.load(f)
                if isinstance(saved, dict):
                    data.update(saved)  # 以已保存內容覆蓋預設
        except Exception:
            pass
    return JSONResponse(data)

@app.post("/notices")
async def save_notices(
    request: Request,
    session_id: str = Form(...),
    key: str = Form(...),           # "operation" 或 "bonding"
    text: str = Form(...),
):
    """儲存單一類別的注意事項；只有 is_editor 的來源 IP 可寫入。"""
    if not is_editor(request):      # 你現有的 is_editor() 會判斷來源 IP & 環境
        return JSONResponse({"error": "forbidden"}, status_code=403)
    if key not in ("operation", "bonding"):
        return JSONResponse({"error": "invalid key"}, status_code=400)

    sess_dir = os.path.join(UPLOAD_DIR, session_id)
    os.makedirs(sess_dir, exist_ok=True)
    path = os.path.join(sess_dir, "notices.json")

    data = {}
    if os.path.exists(path):
        try:
            with open(path, "r", encoding="utf-8") as f:
                data = json.load(f) or {}
        except Exception:
            data = {}

    data[key] = text
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

    return JSONResponse({"ok": True})
